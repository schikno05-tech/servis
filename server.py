"""
Сервис расчёта сдельной смены — версия для хостинга.
• База: SQLite локально, PostgreSQL на сервере (через DATABASE_URL).
• Вход по паролю: роль "brigade" (бригадир) и "admin" (администратор).
Запуск локально:  uvicorn server:app --reload --host 0.0.0.0 --port 8000
"""
import os, datetime, io, csv, secrets, hashlib, urllib.request, urllib.parse
from fastapi import FastAPI, HTTPException, Response, Depends, Cookie
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
from db import get_conn, IS_PG, pk

BASE = os.path.dirname(__file__)
app = FastAPI(title="Расчёт смены")

SPREADSHEET_KEY = os.environ.get("SPREADSHEET_KEY", "1mzPh12X2MCdaWL-pXMTxVAHZWVtpKus0D1UnOCG_lMc")
VITRINA_GID     = os.environ.get("VITRINA_GID", "879590353")
PASS_BRIGADE = os.environ.get("PASS_BRIGADE", "smena2026")
PASS_ADMIN   = os.environ.get("PASS_ADMIN", "admin2026")
SESSIONS = {}

def _migrate_attendance_rate():
    """Добавляет колонку rate в attendance, если её ещё нет (для старых баз).
    Отдельной транзакцией, чтобы ошибка 'колонка уже есть' не ломала инициализацию."""
    try:
        with get_conn() as c:
            c.execute("ALTER TABLE attendance ADD COLUMN rate REAL")
    except Exception:
        pass  # колонка уже есть — это нормально

def init_db():
    with get_conn() as c:
        c.execute(f"CREATE TABLE IF NOT EXISTS shift(id {pk()}, name TEXT NOT NULL)")
        c.execute(f"""CREATE TABLE IF NOT EXISTS worker(
            id {pk()}, name TEXT NOT NULL, rate REAL NOT NULL DEFAULT 0,
            shift_id INTEGER REFERENCES shift(id), active INTEGER NOT NULL DEFAULT 1)""")
        c.execute("CREATE TABLE IF NOT EXISTS product(article TEXT PRIMARY KEY, price REAL NOT NULL DEFAULT 0)")
        c.execute(f"""CREATE TABLE IF NOT EXISTS workday(
            id {pk()}, date TEXT NOT NULL, shift_id INTEGER REFERENCES shift(id),
            lunch_min INTEGER NOT NULL DEFAULT 0, closed INTEGER NOT NULL DEFAULT 0,
            UNIQUE(date, shift_id))""")
        c.execute(f"""CREATE TABLE IF NOT EXISTS attendance(
            id {pk()}, workday_id INTEGER REFERENCES workday(id) ON DELETE CASCADE,
            worker_id INTEGER REFERENCES worker(id), start_ts TEXT, end_ts TEXT,
            rate REAL,
            UNIQUE(workday_id, worker_id))""")
        c.execute(f"""CREATE TABLE IF NOT EXISTS production(
            id {pk()}, workday_id INTEGER REFERENCES workday(id) ON DELETE CASCADE,
            article TEXT, qty REAL NOT NULL DEFAULT 0, ts TEXT)""")
        if not c.execute("SELECT 1 FROM shift").fetchone():
            c.execute("INSERT INTO shift(name) VALUES('Смена 1')")
            c.execute("INSERT INTO shift(name) VALUES('Смена 2')")
    _migrate_attendance_rate()
init_db()

def current_role(session: Optional[str] = Cookie(default=None)):
    return SESSIONS.get(session) if session else None
def require_login(role = Depends(current_role)):
    if not role: raise HTTPException(401, "Требуется вход")
    return role
def require_admin(role = Depends(current_role)):
    if role != "admin": raise HTTPException(403, "Только для администратора")
    return role

class LoginIn(BaseModel):
    password: str
@app.post("/api/login")
def login(body: LoginIn, response: Response):
    if body.password == PASS_ADMIN: role="admin"
    elif body.password == PASS_BRIGADE: role="brigade"
    else: raise HTTPException(401, "Неверный пароль")
    token=secrets.token_urlsafe(24); SESSIONS[token]=role
    response.set_cookie("session", token, httponly=True, max_age=60*60*24*30, samesite="lax")
    return {"role": role}
@app.post("/api/logout")
def logout(response: Response, session: Optional[str] = Cookie(default=None)):
    if session in SESSIONS: del SESSIONS[session]
    response.delete_cookie("session"); return {"ok": True}
@app.get("/api/me")
def me(role = Depends(current_role)):
    return {"role": role}

def worked_hours(start_ts, end_ts, lunch_min, now=None):
    if not start_ts: return 0.0
    if end_ts and str(end_ts).endswith("T00:00"): end_ts=None
    start=datetime.datetime.fromisoformat(start_ts)
    end=datetime.datetime.fromisoformat(end_ts) if end_ts else (now or datetime.datetime.now())
    sec=(end-start).total_seconds()-(lunch_min or 0)*60
    return max(0.0, round(sec/3600,3))

def calc(c, workday_id, now=None):
    wd=c.execute("SELECT * FROM workday WHERE id=?", (workday_id,)).fetchone()
    if not wd: raise HTTPException(404,"Смена не найдена")
    now=now or datetime.datetime.now()
    output=0.0; items=[]
    for r in c.execute("""SELECT pr.article, SUM(pr.qty) qty, p.price
                          FROM production pr LEFT JOIN product p ON p.article=pr.article
                          WHERE pr.workday_id=? GROUP BY pr.article, p.price""",(workday_id,)):
        price=r["price"] or 0; line=(r["qty"] or 0)*price; output+=line
        items.append({"article":r["article"],"qty":r["qty"],"price":price,"sum":round(line,2)})
    workers=[]; total_hours=0.0; total_base=0.0
    for a in c.execute("""SELECT a.start_ts,a.end_ts,a.worker_id,a.rate AS att_rate,w.name,w.rate AS worker_rate
                          FROM attendance a JOIN worker w ON w.id=a.worker_id
                          WHERE a.workday_id=?""",(workday_id,)):
        # ставка фиксируется в смене (att_rate). Если пусто (старые записи) — берём текущую из карточки.
        rate = a["att_rate"] if a["att_rate"] is not None else a["worker_rate"]
        h=worked_hours(a["start_ts"],a["end_ts"],wd["lunch_min"],now); base=h*rate
        total_hours+=h; total_base+=base
        workers.append({"worker_id":a["worker_id"],"name":a["name"],"rate":rate,
            "start_ts":a["start_ts"],"end_ts":a["end_ts"],"hours":round(h,2),
            "on_shift":a["start_ts"] is not None and a["end_ts"] is None,"base":round(base,2)})
    bonus_fund=output-total_base
    bonus_rate=(bonus_fund/total_hours) if total_hours>0 else 0.0
    for w in workers:
        h=w["hours"]; w["bonus"]=round(bonus_rate*h,2); w["total"]=round(w["base"]+w["bonus"],2)
    return {"workday_id":workday_id,"date":wd["date"],"shift_id":wd["shift_id"],
            "closed":bool(wd["closed"]),"lunch_min":wd["lunch_min"],
            "output":round(output,2),"total_base":round(total_base,2),
            "bonus_fund":round(bonus_fund,2),"bonus_rate":round(bonus_rate,4),
            "total_hours":round(total_hours,2),"items":items,"workers":workers,
            "now":now.isoformat(timespec="minutes")}

class WorkerIn(BaseModel):
    name:str; rate:float; shift_id:int
class RateIn(BaseModel):
    rate:float
class OpenIn(BaseModel):
    date:str; shift_id:int; lunch_min:int=0
class AttIn(BaseModel):
    worker_id:int; start_ts:Optional[str]=None; end_ts:Optional[str]=None
class ProdIn(BaseModel):
    article:str; qty:float
class QtyIn(BaseModel):
    qty:float

@app.get("/api/shifts")
def shifts(role=Depends(require_login)):
    with get_conn() as c: return [dict(r) for r in c.execute("SELECT * FROM shift")]

@app.get("/api/workers")
def workers(shift_id:Optional[int]=None, role=Depends(require_login)):
    with get_conn() as c:
        if shift_id:
            rows=c.execute("SELECT * FROM worker WHERE active=1 AND shift_id=?",(shift_id,)).fetchall()
        else:
            rows=c.execute("SELECT * FROM worker WHERE active=1").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/workers")
def add_worker(w:WorkerIn, role=Depends(require_admin)):
    with get_conn() as c:
        c.execute("INSERT INTO worker(name,rate,shift_id) VALUES(?,?,?)",(w.name,w.rate,w.shift_id))
        return {"ok":True}

@app.delete("/api/workers/{wid}")
def del_worker(wid:int, role=Depends(require_admin)):
    with get_conn() as c:
        c.execute("UPDATE worker SET active=0 WHERE id=?",(wid,)); return {"ok":True}

@app.patch("/api/workers/{wid}/rate")
def change_worker_rate(wid:int, body:RateIn, role=Depends(require_admin)):
    """Меняет ставку работника. Действует только ВПЕРЁД: прошлые смены хранят
    свою зафиксированную ставку в attendance и не пересчитываются."""
    with get_conn() as c:
        c.execute("UPDATE worker SET rate=? WHERE id=?",(body.rate,wid))
        return {"ok":True}

@app.post("/api/refresh-prices")
def refresh_prices(role=Depends(require_login)):
    url=f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_KEY}/export?format=csv&gid={VITRINA_GID}"
    try:
        req=urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        data=urllib.request.urlopen(req, timeout=20).read().decode("utf-8")
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать таблицу: {e}. Проверьте доступ «по ссылке — просмотр».")
    rows=list(csv.reader(io.StringIO(data))); n=0; skipped=0
    with get_conn() as c:
        for r in rows[1:]:
            if len(r)<2: continue
            art=(r[0] or "").strip()
            raw=str(r[1]).replace("\xa0","").replace(" ","").replace(",",".").strip()
            try: price=float(raw)
            except: skipped+=1; continue
            if art:
                c.execute("""INSERT INTO product(article,price) VALUES(?,?)
                             ON CONFLICT(article) DO UPDATE SET price=excluded.price""",(art,price))
                n+=1
    return {"updated":n,"skipped":skipped}

@app.get("/api/products/search")
def search_products(q:str="", role=Depends(require_login)):
    with get_conn() as c:
        q=q.strip()
        if not q:
            rows=c.execute("SELECT article,price FROM product ORDER BY article LIMIT 20").fetchall()
        else:
            rows=c.execute("SELECT article,price FROM product WHERE article LIKE ? ORDER BY article LIMIT 20",(f"%{q}%",)).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/products/count")
def products_count(role=Depends(require_login)):
    with get_conn() as c:
        return {"count": c.execute("SELECT COUNT(*) n FROM product").fetchone()["n"]}

@app.post("/api/workday/open")
def open_day(o:OpenIn, role=Depends(require_login)):
    with get_conn() as c:
        if IS_PG:
            c.execute("INSERT INTO workday(date,shift_id,lunch_min) VALUES(?,?,?) ON CONFLICT(date,shift_id) DO NOTHING",(o.date,o.shift_id,o.lunch_min))
        else:
            c.execute("INSERT OR IGNORE INTO workday(date,shift_id,lunch_min) VALUES(?,?,?)",(o.date,o.shift_id,o.lunch_min))
        wid=c.execute("SELECT id FROM workday WHERE date=? AND shift_id=?",(o.date,o.shift_id)).fetchone()["id"]
        for w in c.execute("SELECT id, rate FROM worker WHERE shift_id=? AND active=1",(o.shift_id,)).fetchall():
            if IS_PG:
                c.execute("INSERT INTO attendance(workday_id,worker_id,rate) VALUES(?,?,?) ON CONFLICT(workday_id,worker_id) DO NOTHING",(wid,w["id"],w["rate"]))
            else:
                c.execute("INSERT OR IGNORE INTO attendance(workday_id,worker_id,rate) VALUES(?,?,?)",(wid,w["id"],w["rate"]))
        if o.lunch_min: c.execute("UPDATE workday SET lunch_min=? WHERE id=?",(o.lunch_min,wid))
        return {"workday_id":wid}

@app.post("/api/workday/{wid}/attendance")
def set_att(wid:int, a:AttIn, role=Depends(require_login)):
    with get_conn() as c:
        wr=c.execute("SELECT rate FROM worker WHERE id=?",(a.worker_id,)).fetchone()
        rt=wr["rate"] if wr else 0
        # ставка фиксируется при первом попадании в смену; если уже зафиксирована — не трогаем
        c.execute("""INSERT INTO attendance(workday_id,worker_id,start_ts,end_ts,rate) VALUES(?,?,?,?,?)
                     ON CONFLICT(workday_id,worker_id) DO UPDATE SET start_ts=excluded.start_ts,end_ts=excluded.end_ts,
                     rate=COALESCE(attendance.rate, excluded.rate)""",
                  (wid,a.worker_id,a.start_ts,a.end_ts,rt))
        return {"ok":True}

@app.post("/api/workday/{wid}/add-worker/{worker_id}")
def add_worker_to_day(wid:int, worker_id:int, role=Depends(require_login)):
    with get_conn() as c:
        wr=c.execute("SELECT rate FROM worker WHERE id=?",(worker_id,)).fetchone()
        rt=wr["rate"] if wr else 0
        if IS_PG:
            c.execute("INSERT INTO attendance(workday_id,worker_id,rate) VALUES(?,?,?) ON CONFLICT(workday_id,worker_id) DO NOTHING",(wid,worker_id,rt))
        else:
            c.execute("INSERT OR IGNORE INTO attendance(workday_id,worker_id,rate) VALUES(?,?,?)",(wid,worker_id,rt))
        return {"ok":True}

@app.post("/api/workday/{wid}/production")
def add_prod(wid:int, p:ProdIn, role=Depends(require_login)):
    with get_conn() as c:
        c.execute("INSERT INTO production(workday_id,article,qty,ts) VALUES(?,?,?,?)",
                  (wid,p.article,p.qty,datetime.datetime.now().isoformat(timespec="minutes")))
        return {"ok":True}

@app.get("/api/workday/{wid}/production")
def list_prod(wid:int, role=Depends(require_login)):
    with get_conn() as c:
        rows=c.execute("""SELECT pr.id, pr.article, pr.qty, pr.ts, p.price
                          FROM production pr LEFT JOIN product p ON p.article=pr.article
                          WHERE pr.workday_id=? ORDER BY pr.id DESC""",(wid,)).fetchall()
        return [{"id":r["id"],"article":r["article"],"qty":r["qty"],"price":r["price"] or 0,
                 "sum":round((r["qty"] or 0)*(r["price"] or 0),2),"ts":r["ts"]} for r in rows]

@app.delete("/api/production/{pid}")
def del_prod(pid:int, role=Depends(require_login)):
    with get_conn() as c:
        c.execute("DELETE FROM production WHERE id=?",(pid,)); return {"ok":True}

@app.patch("/api/production/{pid}")
def update_prod(pid:int, q:QtyIn, role=Depends(require_login)):
    with get_conn() as c:
        c.execute("UPDATE production SET qty=? WHERE id=?",(q.qty,pid)); return {"ok":True}

@app.get("/api/workday/{wid}/calc")
def get_calc(wid:int, role=Depends(require_login)):
    with get_conn() as c: return calc(c, wid)

@app.post("/api/workday/{wid}/close")
def close_day(wid:int, role=Depends(require_login)):
    with get_conn() as c:
        c.execute("UPDATE workday SET closed=1 WHERE id=?",(wid,)); return {"ok":True}

@app.post("/api/workday/{wid}/reopen")
def reopen_day(wid:int, role=Depends(require_login)):
    with get_conn() as c:
        c.execute("UPDATE workday SET closed=0 WHERE id=?",(wid,)); return {"ok":True}

@app.get("/api/history")
def history(role=Depends(require_login)):
    with get_conn() as c:
        out=[]
        for wd in c.execute("SELECT * FROM workday ORDER BY date DESC, shift_id").fetchall():
            r=calc(c, wd["id"])
            sh=c.execute("SELECT name FROM shift WHERE id=?",(wd["shift_id"],)).fetchone()
            out.append({"workday_id":wd["id"],"date":wd["date"],"shift_id":wd["shift_id"],
                "shift":sh["name"] if sh else "","closed":bool(wd["closed"]),
                "output":r["output"],"total_hours":r["total_hours"],
                "workers":len([w for w in r["workers"] if w["hours"]>0])})
        return out

@app.get("/api/history/{wid}")
def history_detail(wid:int, role=Depends(require_login)):
    with get_conn() as c: return calc(c, wid)

@app.get("/api/summary")
def summary(date_from:str, date_to:str, role=Depends(require_login)):
    with get_conn() as c:
        wds=c.execute("SELECT id FROM workday WHERE date BETWEEN ? AND ?",(date_from,date_to)).fetchall()
        by_worker={}; by_product={}; total_output=0.0
        for wd in wds:
            r=calc(c, wd["id"]); total_output+=r["output"]
            for w in r["workers"]:
                if w["hours"]<=0: continue
                s=by_worker.setdefault(w["name"],{"hours":0,"total":0})
                s["hours"]+=w["hours"]; s["total"]+=w["total"]
            for it in r["items"]:
                p=by_product.setdefault(it["article"],{"qty":0,"sum":0})
                p["qty"]+=it["qty"]; p["sum"]+=it["sum"]
        for s in by_worker.values(): s["hours"]=round(s["hours"],2); s["total"]=round(s["total"],2)
        for p in by_product.values(): p["qty"]=round(p["qty"],2); p["sum"]=round(p["sum"],2)
        return {"date_from":date_from,"date_to":date_to,"total_output":round(total_output,2),
                "by_worker":by_worker,"by_product":by_product}

@app.get("/api/export")
def export(date_from:str, date_to:str, fmt:str="xlsx", role=Depends(require_login)):
    with get_conn() as c:
        wds=c.execute("""SELECT w.id, w.date, w.shift_id, s.name shift_name
                         FROM workday w LEFT JOIN shift s ON s.id=w.shift_id
                         WHERE w.date BETWEEN ? AND ? ORDER BY w.date, w.shift_id""",(date_from,date_to)).fetchall()
        people_rows=[]; worker_tot={}; product_tot={}; shift_rows=[]
        for wd in wds:
            r=calc(c, wd["id"])
            shift_rows.append((wd["date"], wd["shift_name"] or "", r["output"], r["total_hours"]))
            for w in r["workers"]:
                if w["hours"]<=0: continue
                people_rows.append((wd["date"],wd["shift_name"] or "",w["name"],w["hours"],w["base"],w["bonus"],w["total"]))
                t=worker_tot.setdefault(w["name"],{"hours":0,"total":0}); t["hours"]+=w["hours"]; t["total"]+=w["total"]
            for it in r["items"]:
                p=product_tot.setdefault(it["article"],{"qty":0,"sum":0}); p["qty"]+=it["qty"]; p["sum"]+=it["sum"]
    if fmt=="csv":
        buf=io.StringIO(); wr=csv.writer(buf)
        wr.writerow(["ПО ЛЮДЯМ (по дням)"]); wr.writerow(["Дата","Смена","Работник","Часов","Повременно","Премия","Итого"])
        for row in people_rows: wr.writerow(row)
        wr.writerow([]); wr.writerow(["ИТОГО ПО ЛЮДЯМ за период"]); wr.writerow(["Работник","Часов","Заработал"])
        for n,t in sorted(worker_tot.items()): wr.writerow([n,round(t["hours"],2),round(t["total"],2)])
        wr.writerow([]); wr.writerow(["ПО ТОВАРАМ"]); wr.writerow(["Артикул","Кол-во","Сумма"])
        for a,p in sorted(product_tot.items()): wr.writerow([a,round(p["qty"],2),round(p["sum"],2)])
        data=buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(io.BytesIO(data), media_type="text/csv",
            headers={"Content-Disposition":f'attachment; filename="export_{date_from}_{date_to}.csv"'})
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb=Workbook(); HEAD=PatternFill("solid",start_color="1F3B57")
    def head(ws,cols):
        ws.append(cols)
        for i in range(1,len(cols)+1):
            cell=ws.cell(ws.max_row,i); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=HEAD
    ws1=wb.active; ws1.title="По людям (по дням)"
    head(ws1,["Дата","Смена","Работник","Часов","Повременно","Премия","Итого"])
    for row in people_rows: ws1.append(list(row))
    ws1.append([]); ws1.append(["ИТОГО за период по каждому:"]); ws1.cell(ws1.max_row,1).font=Font(bold=True)
    head(ws1,["Работник","","","Часов","","","Заработал"])
    for n,t in sorted(worker_tot.items()): ws1.append([n,"","",round(t["hours"],2),"","",round(t["total"],2)])
    for col,w in {'A':12,'B':10,'C':22,'D':9,'E':12,'F':10,'G':11}.items(): ws1.column_dimensions[col].width=w
    ws2=wb.create_sheet("По товарам"); head(ws2,["Артикул","Кол-во","Сумма"])
    for a,p in sorted(product_tot.items()): ws2.append([a,round(p["qty"],2),round(p["sum"],2)])
    ws2.column_dimensions['A'].width=34; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=12
    ws3=wb.create_sheet("По сменам"); head(ws3,["Дата","Смена","Выработка","Часов всего"])
    for row in shift_rows: ws3.append([row[0],row[1],round(row[2],2),round(row[3],2)])
    for col,w in {'A':12,'B':12,'C':12,'D':12}.items(): ws3.column_dimensions[col].width=w
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f'attachment; filename="export_{date_from}_{date_to}.xlsx"'})

class PassIn(BaseModel):
    brigade: Optional[str]=None
    admin: Optional[str]=None
@app.post("/api/admin/passwords")
def change_passwords(body:PassIn, role=Depends(require_admin)):
    global PASS_BRIGADE, PASS_ADMIN
    if body.brigade: PASS_BRIGADE=body.brigade
    if body.admin: PASS_ADMIN=body.admin
    return {"ok":True}

@app.get("/", response_class=HTMLResponse)
def index(role = Depends(current_role)):
    return FileResponse(os.path.join(BASE, "index.html" if role else "login.html"))
@app.get("/login", response_class=HTMLResponse)
def login_page():
    return FileResponse(os.path.join(BASE, "login.html"))
